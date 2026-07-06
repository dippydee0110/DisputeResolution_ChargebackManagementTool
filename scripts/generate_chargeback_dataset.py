#!/usr/bin/env python3
"""
Generate sample chargeback dataset in Excel format.
Creates realistic chargeback records for testing dispute detection system.
"""

import os
import json
from datetime import datetime, timedelta
import random

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("⚠️  openpyxl not installed. Installing...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_sample_data():
    """Generate realistic chargeback record samples."""
    merchants = [
        "TechStore Pro", "GlobalRetail Inc", "FashionHub Online",
        "ElectroWorld", "GroceryChain Corp", "HealthPlus Pharmacy",
        "TravelBookings Ltd", "DigitalServices Co"
    ]
    
    reason_codes = [
        ("4837", "Fraud — Cardholder Does Not Recognize"),
        ("4855", "Goods/Services Not Provided"),
        ("4856", "Fraudulent Transfer"),
        ("4863", "Cardholder Does Not Recognize"),
        ("4871", "Chip Liability"),
        ("4901", "Cancelled Recurring Transaction")
    ]
    
    statuses = ["Opened", "Under Investigation", "Representment Ready", "Escalated"]
    
    records = []
    base_date = datetime.now() - timedelta(days=60)
    
    for i in range(15):
        trans_date = base_date + timedelta(days=random.randint(0, 60))
        complaint_date = trans_date + timedelta(days=random.randint(1, 10))
        
        reason_code, reason_desc = random.choice(reason_codes)
        
        records.append({
            "Case ID": f"CASE{1000 + i}",
            "Transaction Date": trans_date.strftime("%Y-%m-%d"),
            "Complaint Date": complaint_date.strftime("%Y-%m-%d"),
            "Days to Complaint": (complaint_date - trans_date).days,
            "Merchant": random.choice(merchants),
            "Amount": round(random.uniform(50, 2000), 2),
            "Currency": "USD",
            "Reason Code": reason_code,
            "Reason Description": reason_desc,
            "Card Network": random.choice(["Visa", "Mastercard"]),
            "Status": random.choice(statuses),
            "Dispute Risk Score": round(random.uniform(0.3, 0.95), 2),
            "Early Signals Detected": random.randint(0, 5),
            "Friendly Fraud Indicator": random.choice(["Yes", "No", "Maybe"]),
            "Has Refund Request": random.choice(["Yes", "No"]),
            "Shipping Address Match": random.choice(["Yes", "No", "Partial"]),
            "AVS Result": random.choice(["Match", "No Match", "Unavailable"]),
            "Notes": random.choice([
                "Cardholder claims unauthorized",
                "Item not delivered",
                "Quality complaint",
                "Duplicate charge",
                "Multiple complaints from same cardholder"
            ])
        })
    
    return records

def create_excel_workbook(records):
    """Create formatted Excel workbook with chargeback data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chargebacks"
    
    # Define styles
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    high_risk_fill = PatternFill(start_color="fecaca", end_color="fecaca", fill_type="solid")
    medium_risk_fill = PatternFill(start_color="fed7aa", end_color="fed7aa", fill_type="solid")
    low_risk_fill = PatternFill(start_color="bbf7d0", end_color="bbf7d0", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column headers
    headers = [
        "Case ID", "Transaction Date", "Complaint Date", "Days to Complaint",
        "Merchant", "Amount", "Currency", "Reason Code", "Reason Description",
        "Card Network", "Status", "Dispute Risk Score", "Early Signals", 
        "Friendly Fraud?", "Refund Request?", "Shipping Match", "AVS Result", "Notes"
    ]
    
    ws.append(headers)
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data rows
    for record in records:
        row = [
            record["Case ID"],
            record["Transaction Date"],
            record["Complaint Date"],
            record["Days to Complaint"],
            record["Merchant"],
            record["Amount"],
            record["Currency"],
            record["Reason Code"],
            record["Reason Description"],
            record["Card Network"],
            record["Status"],
            record["Dispute Risk Score"],
            record["Early Signals Detected"],
            record["Friendly Fraud Indicator"],
            record["Has Refund Request"],
            record["Shipping Address Match"],
            record["AVS Result"],
            record["Notes"]
        ]
        ws.append(row)
    
    # Format data rows and apply conditional coloring
    for idx, record in enumerate(records, start=2):
        risk_score = record["Dispute Risk Score"]
        
        # Determine fill color based on risk
        if risk_score >= 0.7:
            fill = high_risk_fill
        elif risk_score >= 0.5:
            fill = medium_risk_fill
        else:
            fill = low_risk_fill
        
        for cell in ws[idx]:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if cell.column == 12:  # Risk Score column
                cell.fill = fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column in [2, 3, 6]:  # Date and Amount columns
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    column_widths = {
        'A': 12, 'B': 15, 'C': 15, 'D': 18, 'E': 20, 'F': 12, 'G': 10,
        'H': 12, 'I': 30, 'J': 12, 'K': 18, 'L': 18, 'M': 14, 'N': 16,
        'O': 16, 'P': 15, 'Q': 12, 'R': 25
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    return wb

def main():
    """Main execution."""
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    OUTPUT_DIR = os.path.join(BASE_DIR, "output")
    
    print("[*] Generating sample chargeback dataset...")
    records = generate_sample_data()
    
    print(f"[OK] Generated {len(records)} sample chargeback records")
    
    print("[*] Creating Excel workbook...")
    wb = create_excel_workbook(records)
    
    # Save workbook
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_file = os.path.join(OUTPUT_DIR, "chargeback_records.xlsx")
    wb.save(output_file)
    
    print(f"[OK] Excel file saved: {output_file}")
    print(f"\n[INFO] Sample Data Summary:")
    print(f"   - Total Records: {len(records)}")
    print(f"   - Avg Risk Score: {sum(r['Dispute Risk Score'] for r in records)/len(records):.2f}")
    print(f"   - High Risk Cases: {sum(1 for r in records if r['Dispute Risk Score'] >= 0.7)}")

if __name__ == "__main__":
    main()
